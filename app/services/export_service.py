"""
Servicio de exportación a Excel y PDF
"""
from typing import List
from datetime import datetime
import io
from sqlalchemy.orm import Session
from app.models.checklist_entry import ChecklistEntry

class ExportService:
    """Servicio para exportar datos a diferentes formatos"""
    
    def export_to_excel(self, entries: List[ChecklistEntry]) -> bytes:
        """
        Exportar entradas a Excel usando openpyxl
        """
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Historial Checklist"
            
            # Encabezados
            headers = [
                "ID", "Fecha/Hora", "Área", "Turno", "Etapa Protocolo",
                "Cumple", "Observaciones", "Usuario"
            ]
            ws.append(headers)
            
            # Estilo de encabezados
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # Datos
            for entry in entries:
                ws.append([
                    entry.id,
                    entry.fecha_hora.strftime("%Y-%m-%d %H:%M:%S"),
                    entry.area,
                    entry.turno,
                    entry.protocolo_etapa,
                    "Sí" if entry.cumple else "No",
                    entry.observaciones or "",
                    str(entry.user_id)
                ])
            
            # Ajustar anchos de columna
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Guardar en BytesIO
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            return output.read()
            
        except ImportError:
            raise Exception("openpyxl no está instalado. Instala con: pip install openpyxl")
    
    def export_to_csv(self, entries: List[ChecklistEntry]) -> str:
        """
        Exportar entradas a CSV
        """
        import csv
        import io
        
        output = io.StringIO()
        writer = csv.writer(output)
        
        # Encabezados
        writer.writerow([
            "ID", "Fecha/Hora", "Área", "Turno", "Etapa Protocolo",
            "Cumple", "Observaciones", "Usuario"
        ])
        
        # Datos
        for entry in entries:
            writer.writerow([
                entry.id,
                entry.fecha_hora.strftime("%Y-%m-%d %H:%M:%S"),
                entry.area,
                entry.turno,
                entry.protocolo_etapa,
                "Sí" if entry.cumple else "No",
                entry.observaciones or "",
                str(entry.user_id)
            ])
        
        return output.getvalue()
    
    def export_report_to_pdf(self, summary: dict) -> bytes:
        """
        Exportar reporte de cumplimiento a PDF usando reportlab
        """
        try:
            from reportlab.lib.pagesizes import letter, A4
            from reportlab.lib import colors
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.units import inch
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
            
            buffer = io.BytesIO()
            doc = SimpleDocTemplate(buffer, pagesize=letter)
            elements = []
            styles = getSampleStyleSheet()
            
            # Título
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=24,
                textColor=colors.HexColor('#4472C4'),
                spaceAfter=30,
                alignment=1  # Center
            )
            elements.append(Paragraph("Reporte de Cumplimiento - MedCheck", title_style))
            elements.append(Spacer(1, 0.3*inch))
            
            # Fecha del reporte
            fecha_reporte = datetime.now().strftime("%d/%m/%Y %H:%M")
            elements.append(Paragraph(f"Generado: {fecha_reporte}", styles['Normal']))
            elements.append(Spacer(1, 0.3*inch))
            
            # Resumen general
            elements.append(Paragraph("Resumen General", styles['Heading2']))
            general_data = [
                ['Métrica', 'Valor'],
                ['Total de Registros', str(summary.get('total_registros', 0))],
                ['Registros que Cumplen', str(summary.get('total_cumple', 0))],
                ['Porcentaje de Cumplimiento', f"{summary.get('porcentaje_cumplimiento', 0):.1f}%"]
            ]
            
            general_table = Table(general_data, colWidths=[3*inch, 2*inch])
            general_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            elements.append(general_table)
            elements.append(Spacer(1, 0.3*inch))
            
            # Cumplimiento por Área
            if 'por_area' in summary and summary['por_area']:
                elements.append(Paragraph("Cumplimiento por Área", styles['Heading2']))
                area_data = [['Área', 'Total', 'Cumple', 'Porcentaje']]
                for area, stats in summary['por_area'].items():
                    area_data.append([
                        area,
                        str(stats['total']),
                        str(stats['cumple']),
                        f"{stats['porcentaje']:.1f}%"
                    ])
                
                area_table = Table(area_data, colWidths=[2*inch, 1*inch, 1*inch, 1.5*inch])
                area_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 10),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                    ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black)
                ]))
                elements.append(area_table)
                elements.append(Spacer(1, 0.3*inch))
            
            # Construir PDF
            doc.build(elements)
            buffer.seek(0)
            return buffer.read()
            
        except ImportError:
            raise Exception("reportlab no está instalado. Instala con: pip install reportlab")
